import sys
import cv2
from mtcnn import MTCNN
import pandas as pd
from datetime import datetime
import os
from PyQt5 import uic
from PyQt5.uic import loadUi
from PyQt5 import QtCore, QtWidgets, QtGui
from PyQt5.QtGui import QCursor, QStandardItemModel, QStandardItem, QIcon, QPixmap
from PyQt5.QtMultimedia import QSound
import openpyxl
from PIL import Image
import numpy as np
from PyQt5.QtCore import QTimer, QDateTime, QFile, QTextStream, QSize, QThread, pyqtSignal, QMetaObject, Qt, Q_ARG
from PyQt5.QtWidgets import QMessageBox, QDialog, QMainWindow, QApplication, QLineEdit, QFileDialog, QLabel, QSizePolicy
import sqlite3
from threading import Thread
from random import randint
from qimage2ndarray import array2qimage
import zipfile
import dlib
import tensorflow as tf
import pickle
from keras.models import load_model
import hashlib


class FaceRecognitionThread(QThread):
    recognition_complete = pyqtSignal(object)
    display_scrollArea = pyqtSignal(bool)
    add_image = pyqtSignal(QtGui.QImage)
    error_occurred = pyqtSignal(str)
    trigger_sound = pyqtSignal(bool, int)
    def __init__(self, saved_encodings, ids, img, threshold, work_dir, sp, facerec, parent=None):
        super().__init__(parent)
        self.saved_encodings = saved_encodings
        self.ids = ids
        self.img = img
        self.threshold = threshold
        self.work_dir = work_dir
        self.font = cv2.FONT_HERSHEY_SIMPLEX
        self.last_unknown_face = None
        self.last_unknown_face_2 = None
        self.dataframes = []
        self.sp = sp
        self.facerec = facerec
        self.last_unknown_face_time = datetime.now()
        self.last_unknown_face_time_2 = datetime.now()

    def run(self):
        face_cascade = cv2.CascadeClassifier(self.work_dir+"/haarcascade_frontalface_default.xml")
        self.color_img = cv2.cvtColor(self.img, cv2.COLOR_BGR2RGB)
        gray = cv2.cvtColor(self.img, cv2.COLOR_BGR2RGB)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)
        if len(faces) == 0:
            # Return empty result if no faces detected
            result = {'img': self.img}
            self.recognition_complete.emit(result)
            return

        for (x, y, w, h) in faces:
            now = datetime.now()
            date = now.strftime("%d_%m_%Y")
            time = now.strftime("%H_%M_%S")
            self.label = None
            face = dlib.rectangle(x, y, x + w, y + h)
            shape = self.sp(gray, face)
            face_encoding = self.facerec.compute_face_descriptor(gray, shape)
            matches = []
            for saved_encoding in self.saved_encodings:
                match = np.linalg.norm(np.array(saved_encoding) - np.array(face_encoding))
                matches.append(match)
            match_index = np.argmin(matches)
            ids = self.ids[match_index]
            if round((1 - matches[match_index]) * 100) > self.threshold:
                if str(self.ids[match_index]).startswith('101'):
                    confidence = "  {0}%".format(round((1 - matches[match_index]) * 100))
                    self.label = "Suspect"
                    color = (255, 0, 0)
                    self.display_scrollArea.emit(True)
                    if self.last_unknown_face is None or (
                            now - self.last_unknown_face_time).seconds >= 10:
                        self.last_unknown_face_2 = self.color_img[y:y + h, x:x + w]
                        self.last_unknown_face_time = now
                        # Save unknown face to a file
                        path = os.path.join(os.getcwd(), "unknown_faces")
                        if not os.path.exists(path):
                            os.makedirs(path)
                        cv2.imwrite(os.path.join(path, f"Suspect-{date}-{time}.jpg"),
                                    self.last_unknown_face_2)
                        try:
                            test = cv2.cvtColor(self.last_unknown_face_2, cv2.COLOR_BGR2RGB)
                        except cv2.error:
                            self.error_occurred.emit("Webcam is not working properly.")
                        unknown_image = cv2.cvtColor(self.last_unknown_face_2, cv2.COLOR_BGR2RGB)
                        q_unknown_image = array2qimage(unknown_image)

                        # Scale the image to fit the QLabel
                        scaled_image = q_unknown_image.scaled(300, 300, QtCore.Qt.KeepAspectRatio)

                        # Emit signal with scaled image
                        self.add_image.emit(scaled_image)
                        self.trigger_sound.emit(True, 2)
                        data = pd.DataFrame(
                            {"Label": "Suspect", "Confidence": [confidence], "Date": [date], "Time": [time],
                             "ID": [ids]})
                        self.dataframes.append(data)

                else:
                    confidence = "  {0}%".format(round((1 - matches[match_index]) * 100))
                    self.label = "Congregation member"
                    color = (0, 255, 0)
                    self.last_unknown_face = None
                    self.last_unknown_face_2 = None
                    data = pd.DataFrame(
                        {"Label": "Unknown", "Confidence": [confidence], "Date": [date], "Time": [time], "ID": [ids]})
                    self.dataframes.append(data)
            else:
                self.label = "Unknown"
                confidence = "  {0}%".format(round((1 - matches[match_index]) * 100))
                color = (255, 0, 0)
                self.display_scrollArea.emit(True)
                if self.last_unknown_face_2 is None or (
                        now - self.last_unknown_face_time_2).seconds >= 10:
                    self.last_unknown_face_2 = self.color_img[y:y + h, x:x + w]
                    self.last_unknown_face_time_2 = now
                    # Save unknown face to a file
                    path = os.path.join(os.getcwd(), "unknown_faces")
                    if not os.path.exists(path):
                        os.makedirs(path)
                    cv2.imwrite(os.path.join(path, f"Unknown-{date}-{time}.jpg"), self.last_unknown_face_2)
                    try:
                        test = cv2.cvtColor(self.last_unknown_face_2, cv2.COLOR_BGR2RGB)
                    except cv2.error:
                        self.error_occurred.emit("Webcam is not working properly.")
                    unknown_image = cv2.cvtColor(self.last_unknown_face_2, cv2.COLOR_BGR2RGB)
                    q_unknown_image = array2qimage(unknown_image)

                    # Scale the image to fit the QLabel
                    scaled_image = q_unknown_image.scaled(300, 300, QtCore.Qt.KeepAspectRatio)

                    # Emit signal with scaled image
                    self.add_image.emit(scaled_image)
                    self.trigger_sound.emit(True, 1)
                    data = pd.DataFrame(
                        {"Label": "Unknown", "Confidence": [confidence], "Date": [date], "Time": [time], "ID": ["-"]})
                    self.dataframes.append(data)

            self.last_unknown_face_time = datetime.now()
            self.last_unknown_face_time_2 = datetime.now()
            (x, y, w, h) = face.left(), face.top(), face.width(), face.height()
            cv2.rectangle(self.img, (x, y), (x + w, y + h), color, 2)
            cv2.putText(self.img, self.label, (x + 5, y - 5), self.font, 1, (255, 255, 255), 2)
            cv2.putText(self.img, str(confidence), (x + 5, y + h - 5), self.font, 1,
                        (255, 255, 255), 2)

            result = {'face': face, 'dataframes':self.dataframes, 'img': self.img}


        if len(self.dataframes) > 0:
            self.write_dataframe_thread = WriteDataFrameThread(self.dataframes, "face_log.xlsx")
            self.write_dataframe_thread.data_list = self.dataframes
            self.write_dataframe_thread.start()
            self.dataframes = []

        self.recognition_complete.emit(result)

class WebcamThreadCollect(QThread):
    initialized = pyqtSignal(bool)

    def run(self):
        # if self.stop_thread:
        #     return
        cap = cv2.VideoCapture(0)
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)
        self.initialized.emit(cap.isOpened())
        self.cap = cap

class WebcamThreadRecog(QThread):
    initialized = pyqtSignal(bool)

    def run(self):
        #while not self.stop_signal.emit():
        cam = cv2.VideoCapture(0)
        cam.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
        cam.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)
        self.initialized.emit(cam.isOpened())
        self.cam = cam

class WriteDataFrameThread(QtCore.QThread):
    def __init__(self, data_list, file_path):
        QtCore.QThread.__init__(self)
        self.data_list = data_list
        self.file_path = file_path

    def run(self):
        try:
            book = openpyxl.load_workbook(self.file_path)
        except FileNotFoundError:
            book = openpyxl.Workbook()
            sheet = book.active
            sheet.title = 'face_log'
            book.save(self.file_path)

        writer = pd.ExcelWriter(self.file_path, engine='openpyxl')
        writer.book = book

        try:
            writer.sheets = {ws.title: ws for ws in book.worksheets}
        except AttributeError:
            pass

        for i in range(len(self.data_list)):
            data = self.data_list[i]
            data.to_excel(writer, sheet_name='face_log', index=False, header=False,
                          startrow=writer.sheets['face_log'].max_row)

        writer.save()
        writer.close()

class LoginForm(QtWidgets.QDialog):
    def __init__(self):
        super(LoginForm, self).__init__()
        loadUi('UI/login_form.ui', self)
        with open("Styles/loginStyle.qss", "r") as f:
            self.stylesheet = f.read()
        self.file = QFile("Styles/loginStyle.qss")
        self.file.open(QFile.ReadOnly | QFile.Text)
        self.stream = QTextStream(self.file)
        self.stylesheet = self.stream.readAll()
        self.setStyleSheet(self.stylesheet)
        self.setFixedSize(QSize(415, 345))
        self.path = os.getcwd()
        self.path_icon_1 = QIcon(self.path+"/Styles/icons/user.svg")
        self.path_icon_2 = QIcon(self.path +"/Styles/icons/lock.svg")
        self.usernameLineEdit.addAction(self.path_icon_1, QLineEdit.LeadingPosition)
        self.passwordLineEdit.addAction(self.path_icon_2, QLineEdit.LeadingPosition)
        self.clearButton.clicked.connect(self.clearString)
        self.loginButton.clicked.connect(self.login)
        self.passwordLineEdit.setEchoMode(QLineEdit.Password)

    def clearString(self):
        self.usernameLineEdit.clear()
        self.passwordLineEdit.clear()

    def showMainWindow(self, window):
        window.show()

    def emptyText(self):
        self.usernameLineEdit.setText("")
        self.passwordLineEdit.setText("")

    def login(self):
        username = self.usernameLineEdit.text()
        password = self.passwordLineEdit.text()

        if not username and not password:
            QMessageBox.warning(self, "Warning", "Please enter a username and password.")
            return

        if not username:
            QMessageBox.warning(self, "Warning", "Please enter a username.")
            return

        if not password:
            QMessageBox.warning(self, "Warning", "Please enter a password.")
            return

        conn = sqlite3.connect('users.db')
        c = conn.cursor()
        # Hash the entered password using the same hash function as the one used to hash the passwords in the database
        hashed_password = hashlib.sha256(password.encode('utf-8')).hexdigest()
        c.execute("SELECT * FROM users WHERE username=? AND password_hash=?", (username, hashed_password))
        row = c.fetchone()

        if row is None:
            QMessageBox.warning(self, "Warning", "Incorrect Username or Password")
        else:
            print("Login successful")
            if username == "admin":
                self.mainmenu = FirstMenu()
            elif username == "security":
                self.mainmenu = FirstMenu_2()
            self.usernameLineEdit.setText("")
            self.passwordLineEdit.setText("")
            self.emptyText()
            self.showMainWindow(self.mainmenu)
            widget.hide()


        conn.close()

class ModelLoader(QThread):
    modelLoaded = pyqtSignal(object, object)

    def __init__(self, work_dir):
        super().__init__()
        self.work_dir = work_dir

    def run(self):
        sp = dlib.shape_predictor(self.work_dir + "/trainer/shape_predictor_68_face_landmarks.dat")
        facerec = dlib.face_recognition_model_v1(self.work_dir + "/trainer/dlib_face_recognition_resnet_model_v1.dat")
        self.modelLoaded.emit(sp, facerec)

class FirstMenu(QMainWindow):
    # parameterized constructor
    def __init__(self):
        super().__init__()
        loadUi("UI/test_main_menu.ui", self)
        self.buttonGroupBox_2.setVisible(False)
        self.collectDataBtn.clicked.connect(self.show_CollectData)
        self.lockBtn.clicked.connect(self.backToLoginForm)
        self.lockBtn.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.openRecognitionCamBtn.clicked.connect(self.show_recognizer)
        self.collectDataBtn.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.openRecognitionCamBtn.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.loadExcelButton.clicked.connect(self.show_Load_Excel)
        self.loadExcelButton.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.enlargeBtn.clicked.connect(self.enlarge)
        self.minimizeBtn.clicked.connect(self.showMinimized)
        self.exitBtn.clicked.connect(self.close)
        self.pushLayoutInBtn.clicked.connect(self.layoutIn1)
        self.pushLayoutInBtn.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.pushLayoutInBtn_3.clicked.connect(self.layoutIn2)
        self.pushLayoutInBtn_3.setCursor(QCursor(QtCore.Qt.PointingHandCursor))

        self.collectDataBtn_3.clicked.connect(self.show_CollectData)
        self.lockBtn_3.clicked.connect(self.backToLoginForm)
        self.lockBtn_3.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.openRecognitionCamBtn_3.clicked.connect(self.show_recognizer)
        self.collectDataBtn_3.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.openRecognitionCamBtn_3.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.loadExcelButton_3.clicked.connect(self.show_Load_Excel)
        self.loadExcelButton_3.setCursor(QCursor(QtCore.Qt.PointingHandCursor))

        self.loadExcelButton_3.setToolTip("Log")
        self.openRecognitionCamBtn_3.setToolTip("Start Recognition Cam")
        self.lockBtn_3.setToolTip("Lock")
        self.pushLayoutInBtn_3.setToolTip("Expand")
        self.collectDataBtn_3.setToolTip("Collect Data")

        self.current_page = None
        self.__isDragging = False
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.bottom_right_resizer = QtWidgets.QSizeGrip(self)
        self.bottom_right_resizer.setStyleSheet("background-color: transparent;")
        self.bottom_right_resizer.setGeometry(self.width() - 16, self.height() - 16, 16, 16)
        self.work_dir = os.getcwd()
        self.modelLoader = ModelLoader(self.work_dir)
        self.modelLoader.modelLoaded.connect(self.onModelsLoaded)
        self.modelLoader.start()

    def onModelsLoaded(self, sp, facerec):
        self.sp = sp
        self.facerec = facerec

    def show_error_message(self, message):
        QtWidgets.QMessageBox.critical(self, "Error", message)

    def trigger_warning_sound(self, check, num):
        if num == 1:
            self.current_page.warning_sound.play()
        elif num == 2:
            self.current_page.warning_sound_2.play()




    def layoutIn1(self):
        self.buttonGroupBox.setVisible(False)
        self.buttonGroupBox_2.setVisible(True)

    def layoutIn2(self):
        self.buttonGroupBox.setVisible(True)
        self.buttonGroupBox_2.setVisible(False)


    def resizeEvent(self, event):
        self.bottom_right_resizer.setGeometry(self.width() - 16, self.height() - 16, 16, 16)

    def mousePressEvent(self, event):
        if event.button() == QtCore.Qt.LeftButton:
            self.__isDragging = True
            self.__startPos = event.pos()
            if self.bottom_right_resizer.rect().contains(event.pos()):
                self.start_resizing(event)

    def mouseReleaseEvent(self, event):
        if event.button() == QtCore.Qt.LeftButton:
            self.__isDragging = False
            if hasattr(self, 'resize_origin'):
                del self.resize_origin

    def mouseMoveEvent(self, event):
        if self.__isDragging:
            delta = event.pos() - self.__startPos
            self.move(self.pos() + delta)
        if hasattr(self, 'resize_origin'):
            delta = event.pos() - self.resize_origin
            new_size = self.resize_origin_size + delta
            self.resize(new_size.width(), new_size.height())

    def start_resizing(self, event):
        self.resize_origin = event.pos()
        self.resize_origin_size = QtCore.QSize(self.width(), self.height())

    def mouseDoubleClickEvent(self, event):
        if event.button() == Qt.LeftButton:
            if self.isMaximized():
                self.showNormal()
            else:
                self.showMaximized()

    def enlarge(self):
        if self.isMaximized():
            self.showNormal()
        else:
            self.showMaximized()

    def backToLoginForm(self):
        self.close()
        if hasattr(self.current_page, "timer"):
            self.current_page.timer.stop()
            self.current_page.cam.release()
        widget.show()

    def on_webcam_initializedCollect(self, success):
        if success:
            self.current_page.cap = self.current_page.thread.cap
        else:
            print("Webcam initialization failed")
        self.current_page.timer = QtCore.QTimer(self)
        self.current_page.stop_button.setEnabled(True)
        self.current_page.timer.timeout.connect(self.update_webcam_collect)
        self.current_page.minW = 0
        self.current_page.minH = 0
        self.current_page.timer.start(30)

    def on_webcam_initializedRecog(self, success):
        if success:
            self.current_page.cam = self.current_page.webcam_thread_recog.cam
        else:
            print("Webcam initialization failed")
        self.current_page.timer = QtCore.QTimer(self)
        self.current_page.stopRecordBtn.setEnabled(True)
        self.current_page.cam.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
        self.current_page.cam.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)
        self.current_page.minW = int(0.1 * self.current_page.cam.get(3))
        self.current_page.minH = int(0.1 * self.current_page.cam.get(4))
        self.current_page.timer.timeout.connect(self.update_frame_recog)
        self.current_page.timer.start(30)

    def show_CollectData(self):
        if self.openRecognitionCamBtn.setDisabled(False):
            self.gotoFirstMenu_recog()
        self.openRecognitionCamBtn.setDisabled(False)
        self.loadExcelButton.setDisabled(False)
        self.openRecognitionCamBtn_3.setDisabled(False)
        self.loadExcelButton_3.setDisabled(False)
        self.collectDataBtn.setDisabled(True)
        self.collectDataBtn_3.setDisabled(True)
        self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
        self.setMinimumSize(0, 0)
        self.resize(1050, 800)
        if self.current_page:
            self.current_page.hide()
        self.current_page = loadUi('UI/Collect_Data.ui')
        self.content_layout.addWidget(self.current_page)

        with open("Styles/recorderStyle.qss", "r") as f:
            self.stylesheet = f.read()
        self.current_page.file = QFile("Styles/recorderStyle.qss")
        self.current_page.file.open(QFile.ReadOnly | QFile.Text)
        self.current_page.stream = QTextStream(self.current_page.file)
        self.current_page.stylesheet = self.current_page.stream.readAll()
        self.current_page.setStyleSheet(self.current_page.stylesheet)

        # Initialize MTCNN detector
        self.current_page.detector = MTCNN()
        self.current_page.counter = 0
        self.current_page.all_ids = set()

        # Retrieving all the prefixes of the images in the current directory
        for file in os.listdir(os.path.join(os.getcwd(), "dataset")):
            if file.endswith(".jpg"):
                self.current_page.all_ids.add(file.split("_")[0])

        # Connect signals
        self.current_page.start_button.clicked.connect(self.start_webcam_collect)
        self.current_page.stop_button.clicked.connect(self.stop_webcam_collect)
        self.current_page.backBtn.clicked.connect(self.gotoFirstMenu_collect)
        self.current_page.trainBtn.clicked.connect(self.start_Train)
        self.current_page.start_button.setIconSize(QSize(int(self.current_page.start_button.width() / 100 * 50),
                                                      int(self.current_page.start_button.height() / 100 * 50)))
        self.current_page.stop_button.setIconSize(QSize(int(self.current_page.stop_button.width() / 100 * 50),
                                                          int(self.current_page.stop_button.height() / 100 * 50)))
        self.current_page.trainBtn.setIconSize(QSize(int(self.current_page.trainBtn.width() / 100 * 50),
                                                     int(self.current_page.trainBtn.height() / 100 * 50)))
        self.current_page.backBtn.setIconSize(QSize(int(self.current_page.backBtn.width() / 100 * 50),
                                                    int(self.current_page.backBtn.height() / 100 * 50)))
        self.current_page.start_button.setCursor(Qt.PointingHandCursor)
        self.current_page.stop_button.setCursor(Qt.PointingHandCursor)
        self.current_page.backBtn.setCursor(Qt.PointingHandCursor)
        self.current_page.trainBtn.setCursor(Qt.PointingHandCursor)
        self.current_page.start_button.setToolTip("Start Recording")
        self.current_page.stop_button.setToolTip("Stop Recording")
        self.current_page.trainBtn.setToolTip("Train Model")
        self.current_page.backBtn.setVisible(False)
        self.current_page.stop_button.setEnabled(False)
        self.current_page.trainBtn.setEnabled(True)

    def start_Train(self):
        self.setDisabled(True)
        trainer = Train_MainWindow(self)
        trainer.exec()

    def stop_webcam_collect(self):
        # Release the VideoCapture object
        self.current_page.cap.release()

        # Stop the QTimer
        self.current_page.timer.stop()

        # Enable buttons and inputs
        self.openRecognitionCamBtn.setEnabled(True)
        self.loadExcelButton.setEnabled(True)
        self.current_page.id_edit.setEnabled(True)
        self.current_page.start_button.setEnabled(True)
        self.current_page.stop_button.setEnabled(False)
        self.current_page.trainBtn.setEnabled(True)
        self.current_page.backBtn.setEnabled(True)

        # Reset the minimum size of the window
        self.setMinimumSize(1050, 800)

        # Reset the size policy of the window
        self.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.MinimumExpanding)

        # Resize the window
        self.resize(1175, 830)

    def gotoFirstMenu_collect(self):
        self.setMinimumSize(1050, 800)
        self.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.MinimumExpanding)
        self.resize(1050, 800)
        self.collectDataBtn.setDisabled(False)
        if self.current_page:
            self.current_page.hide()
        for i in reversed(range(self.content_layout.count())):
            self.content_layout.itemAt(i).widget().setParent(None)
        self.current_page = loadUi('UI/empty_menu.ui')
        self.content_layout.addWidget(self.current_page)

    def start_webcam_collect(self):
        check = self.test_cam(1)
        if check == 1:
            QMessageBox.warning(None, "Warning", "Webcam not found.")
            return
        id = self.current_page.id_edit.text()
        # validating the input only contains numbers
        if id.isnumeric():
            if id in self.current_page.all_ids:
                app = QApplication.instance()
                if app is None:
                    app = QApplication(sys.argv)
                # Load stylesheet
                with open('Styles/messageBox.qss') as f:
                    style = f.read()
                    QApplication.setStyle('Fusion')
                    app.setStyleSheet(style)
                msgBox = QMessageBox()
                msgBox.setIcon(QMessageBox.Warning)
                msgBox.setText("This ID already exists, please try another one")
                msgBox.setWindowTitle("Warning")
                msgBox.setStandardButtons(QMessageBox.Ok)
                msgBox.exec_()
                print("This ID already exists, please try another one")
                return
            else:
                self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
                self.setMinimumSize(0, 0)
                self.current_page.id = id
                self.current_page.id_edit.setEnabled(False)
                self.current_page.start_button.setEnabled(False)
                self.current_page.backBtn.setEnabled(False)
                self.openRecognitionCamBtn.setEnabled(False)
                self.current_page.trainBtn.setEnabled(False)
                self.loadExcelButton.setEnabled(False)
                # def initialize_webcam():
                #     self.current_page.cap = cv2.VideoCapture(0)
                #     self.current_page.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
                #     self.current_page.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)
                #
                # Thread(target=initialize_webcam).start()
                self.current_page.thread = WebcamThreadCollect(self)
                self.current_page.thread.initialized.connect(lambda success: self.on_webcam_initializedCollect(success))
                self.current_page.thread.start()
        elif id == "":
            msgBox = QMessageBox()
            msgBox.setIcon(QMessageBox.Warning)
            msgBox.setText("Please enter an id")
            msgBox.setWindowTitle("Warning")
            msgBox.setStandardButtons(QMessageBox.Ok)
            msgBox.exec_()
        else:
            app = QApplication.instance()
            if app is None:
                app = QApplication(sys.argv)
            with open('Styles/messageBox.qss') as f:
                style = f.read()
                QApplication.setStyle('Fusion')
                app.setStyleSheet(style)
            msgBox = QMessageBox()
            msgBox.setIcon(QMessageBox.Warning)
            msgBox.setText("Invalid input, Please enter only numbers")
            msgBox.setWindowTitle("Warning")
            msgBox.setStandardButtons(QMessageBox.Ok)
            msgBox.exec_()

    def update_webcam_collect(self):
        if self.current_page.cap is None:
            return
        ret, frame = self.current_page.cap.read()
        if frame is None:
            return
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        if ret:
            # Detect faces in the frame
            faces = self.current_page.detector.detect_faces(frame)

            # Draw bounding boxes around the faces
            for face in faces:
                x, y, width, height = face['box']
                cv2.rectangle(frame, (x, y), (x + width, y + height), (0, 0, 255), 2)

                # Crop the face from the frame
                face_img = frame[y:y + height, x:x + width]

                # Convert the face to grayscale
                gray_face = cv2.cvtColor(face_img, cv2.COLOR_BGR2RGB)

                # Save the aligned gray face to a file using the id as prefix
                dataset_path = os.path.join(os.getcwd(), "dataset")
                if not os.path.exists(dataset_path):
                    os.mkdir(dataset_path)
                cv2.imwrite(os.path.join(dataset_path, self.current_page.id + "_face" + str(self.current_page.counter) + ".jpg"), gray_face)
                self.current_page.all_ids.add(self.current_page.id)
                self.current_page.counter += 1
                if self.current_page.counter >= 100:
                    self.stop_webcam_collect()
                    self.current_page.counter = 0
                    self.current_page.id = None
                    reply = QMessageBox.question(self, 'Message',
                                                 "Do you want to start the model training process?", QMessageBox.Yes |
                                                 QMessageBox.No, QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        self.setDisabled(True)
                        trainer = Train_MainWindow(self)
                        trainer.exec()
                    else:
                        pass
                    break

                # Display the frame with bounding boxes
            # self.current_page.webcam_label.setPixmap(QtGui.QPixmap.fromImage(
            #     QtGui.QImage(frame, frame.shape[1], frame.shape[0], QtGui.QImage.Format_RGB888)))

            h, w, ch = frame.shape
            bytes_per_line = ch * w
            # convert_to_Qt_format = QtGui.QImage(self.current_page.img.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
            # p = convert_to_Qt_format.scaled(900, 700)
            self.current_page.webcam_label.setPixmap(QtGui.QPixmap.fromImage(
                QtGui.QImage(frame.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)))

    def show_recognizer(self):
        self.loadExcelButton.setDisabled(False)
        self.collectDataBtn.setDisabled(False)
        self.loadExcelButton_3.setDisabled(False)
        self.collectDataBtn_3.setDisabled(False)
        self.openRecognitionCamBtn.setDisabled(True)
        self.openRecognitionCamBtn_3.setDisabled(True)
        self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
        self.setMinimumSize(0, 0)
        self.resize(1050, 800)
        if self.current_page:
            self.current_page.hide()
        self.current_page = loadUi('UI/Recognizer.ui')
        self.content_layout.addWidget(self.current_page)

        with open("Styles/recognizerStyle.qss", "r") as f:
            self.stylesheet = f.read()
        self.current_page.file = QFile("Styles/recognizerStyle.qss")
        self.current_page.file.open(QFile.ReadOnly | QFile.Text)
        self.current_page.stream = QTextStream(self.current_page.file)
        self.current_page.stylesheet = self.current_page.stream.readAll()
        self.current_page.setStyleSheet(self.current_page.stylesheet)
        self.current_page.webcam_thread_recog = None
        self.current_page.scrollArea.setVisible(False)
        self.current_page.unknown_faces_label.setVisible(False)
        self.current_page.unknown_faces_frame.setVisible(False)
        self.current_page.work_dir = os.getcwd()
        self.current_page.recognizer = cv2.face.LBPHFaceRecognizer_create()
        self.current_page.recognizer.read(self.current_page.work_dir + '/trainer/trainer.yml')
        self.current_page.work_dir = os.getcwd()
        self.current_page.warning_sound = QSound(self.current_page.work_dir + '/Sound/warning.wav')
        self.current_page.warning_sound_2 = QSound(self.current_page.work_dir + '/Sound/warning-sound-2.wav')

        self.current_page.faceCascade = cv2.CascadeClassifier(
            os.path.join(self.current_page.work_dir, "haarcascade_frontalface_default.xml"))
        self.current_page.font = cv2.FONT_HERSHEY_SIMPLEX

        # Define min window size to be recognized as a face
        self.current_page.backBtn.clicked.connect(self.gotoFirstMenu_recog)
        self.current_page.recordBtn.clicked.connect(self.start_webcam_recog)
        self.current_page.stopRecordBtn.clicked.connect(self.stop_webcam_recog)
        self.current_page.clearImage.clicked.connect(self.clear_images)
        self.current_page.backBtn.setIconSize(QSize(int(self.current_page.backBtn.width() / 100 * 50),
                                                         int(self.current_page.backBtn.height() / 100 * 50)))
        self.current_page.recordBtn.setIconSize(QSize(int(self.current_page.recordBtn.width() / 100 * 50),
                                                        int(self.current_page.recordBtn.height() / 100 * 50)))
        self.current_page.stopRecordBtn.setIconSize(QSize(int(self.current_page.stopRecordBtn.width() / 100 * 50),
                                                     int(self.current_page.stopRecordBtn.height() / 100 * 50)))
        self.current_page.backBtn.setVisible(False)
        self.current_page.clearImage.setVisible(False)
        self.current_page.recordBtn.setCursor(Qt.PointingHandCursor)
        self.current_page.stopRecordBtn.setCursor(Qt.PointingHandCursor)
        self.current_page.backBtn.setCursor(Qt.PointingHandCursor)
        self.current_page.recordBtn.setToolTip("Start Recording")
        self.current_page.stopRecordBtn.setToolTip("Stop Recording")
        self.current_page.stopRecordBtn.setEnabled(False)
        self.current_page.last_unknown_face = None
        self.current_page.last_unknown_face_2 = None
        self.current_page.last_unknown_face_time = datetime.now()
        self.current_page.last_unknown_face_time_2 = datetime.now()


    def add_unknown_face_to_layout(self, img_label):
        unknown_label = QtWidgets.QLabel()
        unknown_label.setPixmap(QtGui.QPixmap.fromImage(img_label))
        self.current_page.unknown_faces_layout.addWidget(unknown_label)

    def clear_images(self):
        for i in reversed(range(self.current_page.unknown_faces_layout.count())):
            widget = self.current_page.unknown_faces_layout.itemAt(i).widget()
            if widget is not None:
                widget.setParent(None)

    def gotoFirstMenu_recog(self):
        self.current_page.webcam_thread_recog.quit()
        self.current_page.webcam_thread_recog.wait()
        if self.openRecognitionCamBtn.setDisabled(True):
            if self.current_page.timer and self.current_page.cam:
                self.stop_webcam_recog()
        self.openRecognitionCamBtn.setDisabled(False)
        self.loadExcelButton.setEnabled(True)
        self.collectDataBtn.setEnabled(True)
        self.setMinimumSize(1050, 800)
        self.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.MinimumExpanding)
        self.resize(1050, 800)
        if hasattr(self.current_page, "webcam_thread_recog"):
            if self.current_page.webcam_thread_recog.isRunning():
                self.current_page.webcam_thread_recog.stop_signal.emit()
        if self.current_page:
            self.current_page.hide()
        for i in reversed(range(self.content_layout.count())):
            self.content_layout.itemAt(i).widget().setParent(None)
        self.current_page = loadUi('UI/empty_menu.ui')
        self.content_layout.addWidget(self.current_page)

    def test_cam(self, source):
        self.cap = cv2.VideoCapture(source)
        if not self.cap.isOpened():
            self.cap.release()
            return 1
        return 0

    def start_webcam_recog(self):
        check = self.test_cam(1)
        if check == 1:
            QMessageBox.warning(None, "Warning", "Webcam not found.")
            return
        else:
            self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
            self.setMinimumSize(0, 0)
            self.current_page.recordBtn.setEnabled(False)
            self.current_page.imageLabel.setVisible(True)
            if self.current_page.unknown_faces_layout.count() > 0:
                self.current_page.scrollArea.setVisible(True)
                self.current_page.unknown_faces_label.setVisible(True)
                self.current_page.clearImage.setVisible(True)
                self.current_page.unknown_faces_frame.setVisible(True)
            self.loadExcelButton.setEnabled(False)
            self.collectDataBtn.setEnabled(False)
            self.current_page.webcam_thread_recog = WebcamThreadRecog(self)
            self.current_page.webcam_thread_recog.initialized.connect(
                lambda success: self.on_webcam_initializedRecog(success))
            self.current_page.webcam_thread_recog.start()

    def stop_webcam_recog(self):
        self.resize(1200, 850)
        self.setMinimumSize(1200, 830)
        self.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.MinimumExpanding)
        self.loadExcelButton.setEnabled(True)
        self.collectDataBtn.setEnabled(True)
        self.current_page.recordBtn.setEnabled(True)
        self.current_page.stopRecordBtn.setEnabled(False)
        self.current_page.scrollArea.setVisible(False)
        self.current_page.unknown_faces_label.setVisible(False)
        self.current_page.clearImage.setVisible(False)
        self.current_page.unknown_faces_frame.setVisible(False)
        self.current_page.timer.stop()
        self.current_page.cam.release()
        self.current_page.imageLabel.setPixmap(QtGui.QPixmap())





    def handle_recognition_result(self, result):
        if 'face' not in result:
            img = result['img']
            h, w, ch = img.shape
            bytes_per_line = ch * w
            self.current_page.imageLabel.setPixmap(QtGui.QPixmap.fromImage(
                QtGui.QImage(img.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)))

            self.current_page.unknown_faces_layout.update()
            self.update_frame_recog()  # update the image in the GUI
            return

        img = result['img']
        h, w, ch = img.shape
        bytes_per_line = ch * w
        self.current_page.imageLabel.setPixmap(QtGui.QPixmap.fromImage(
            QtGui.QImage(img.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)))

        self.current_page.unknown_faces_layout.update()
        self.update_frame_recog()  # update the image in the GUI

    def set_ScrollArea_Visible(self):
        self.current_page.scrollArea.setVisible(True)
        self.current_page.unknown_faces_label.setVisible(True)
        self.current_page.clearImage.setVisible(True)
        self.current_page.unknown_faces_frame.setVisible(True)

    def clean_CamRecog(self):
        self.current_page.imageLabel.setPixmap(QtGui.QPixmap())

    def update_frame_recog(self):
        if hasattr(self.current_page, 'recognition_thread'):
            if self.current_page.recognition_thread is not None and self.current_page.recognition_thread.isRunning():
                return
        ret, img = self.current_page.cam.read()
        if not hasattr(self.current_page, 'cam') or img is None:
            return
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        self.current_page.color_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        self.current_page.threshold = 65

        # Load the saved face encodings
        with open(self.current_page.work_dir + '/encodings.pkl', 'rb') as f:
            saved_encodings = pickle.load(f)
        embeddings = saved_encodings["encodings"]
        self.current_page.ids = saved_encodings["ids"]
        self.current_page.recognition_thread = FaceRecognitionThread(embeddings, self.current_page.ids, img,
                                                                     self.current_page.threshold,
                                                                     self.current_page.work_dir, self.sp, self.facerec)
        self.current_page.recognition_thread.recognition_complete.connect(self.handle_recognition_result)
        self.current_page.recognition_thread.add_image.connect(self.add_unknown_face_to_layout)
        self.current_page.recognition_thread.display_scrollArea.connect(self.set_ScrollArea_Visible)
        self.current_page.recognition_thread.trigger_sound.connect(self.trigger_warning_sound)
        self.current_page.recognition_thread.start()



    def show_Load_Excel(self):
        if os.path.isfile("face_log.xlsx"):
            self.openRecognitionCamBtn.setDisabled(False)
            self.collectDataBtn.setDisabled(False)
            self.loadExcelButton.setDisabled(True)
            self.openRecognitionCamBtn_3.setDisabled(False)
            self.collectDataBtn_3.setDisabled(False)
            self.loadExcelButton_3.setDisabled(True)
            self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
            self.setMinimumSize(1050, 800)
            if self.current_page:
                self.current_page.hide()
            self.current_page = loadUi('UI/table.ui')
            self.content_layout.addWidget(self.current_page)
            self.load_excel()
        else:
            QMessageBox.warning(None, "Warning", "Log does not exist yet")
            return

    def load_excel(self):
        # Show a file dialog to choose an Excel file
        #filename, _ = QFileDialog.getOpenFileName(self, "Open Excel File", os.getcwd()+"\Log", "Excel Files (*.xlsx *.xls)")
        filename = os.getcwd() + "/face_log.xlsx"
        if filename:
            # Load the data from the Excel file into a pandas DataFrame
            df = pd.read_excel(filename)

            # Create a new QStandardItemModel and set it to the QTableView
            model = QStandardItemModel(df.shape[0], df.shape[1], self.current_page)
            self.current_page.table.setModel(model)

            # Add the column titles to the model
            for column in range(df.shape[1]):
                title = str(df.columns[column])
                item = QStandardItem(title)
                model.setHorizontalHeaderItem(column, item)

            # Add the data to the model
            for row in range(df.shape[0]):
                for column in range(df.shape[1]):
                    item = QStandardItem(str(df.iloc[row, column]))
                    model.setItem(row, column, item)

class FirstMenu_2(QtWidgets.QMainWindow):
    # parameterized constructor
    def __init__(self):
        super().__init__()
        loadUi("UI/test_main_menu_2.ui", self)
        self.collectDataBtn.setVisible(False)
        self.buttonGroupBox_2.setVisible(False)
        self.collectDataBtn.clicked.connect(self.show_CollectData)
        self.lockBtn.clicked.connect(self.backToLoginForm)
        self.lockBtn.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.openRecognitionCamBtn.clicked.connect(self.show_recognizer)
        self.collectDataBtn.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.openRecognitionCamBtn.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.loadExcelButton.clicked.connect(self.show_Load_Excel)
        self.loadExcelButton.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.enlargeBtn.clicked.connect(self.enlarge)
        self.minimizeBtn.clicked.connect(self.showMinimized)
        self.exitBtn.clicked.connect(self.close)
        self.pushLayoutInBtn.clicked.connect(self.layoutIn1)
        self.pushLayoutInBtn.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.pushLayoutInBtn_3.clicked.connect(self.layoutIn2)
        self.pushLayoutInBtn_3.setCursor(QCursor(QtCore.Qt.PointingHandCursor))

        self.collectDataBtn_3.clicked.connect(self.show_CollectData)
        self.lockBtn_3.clicked.connect(self.backToLoginForm)
        self.lockBtn_3.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.openRecognitionCamBtn_3.clicked.connect(self.show_recognizer)
        self.collectDataBtn_3.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.openRecognitionCamBtn_3.setCursor(QCursor(QtCore.Qt.PointingHandCursor))
        self.loadExcelButton_3.clicked.connect(self.show_Load_Excel)
        self.loadExcelButton_3.setCursor(QCursor(QtCore.Qt.PointingHandCursor))

        self.loadExcelButton_3.setToolTip("Log")
        self.openRecognitionCamBtn_3.setToolTip("Start Recognition Cam")
        self.lockBtn_3.setToolTip("Lock")
        self.pushLayoutInBtn_3.setToolTip("Expand")
        self.collectDataBtn_3.setToolTip("Collect Data")

        self.current_page = None
        self.__isDragging = False
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.bottom_right_resizer = QtWidgets.QSizeGrip(self)
        self.bottom_right_resizer.setStyleSheet("background-color: transparent;")
        self.bottom_right_resizer.setGeometry(self.width() - 16, self.height() - 16, 16, 16)
        self.work_dir = os.getcwd()
        self.modelLoader = ModelLoader(self.work_dir)
        self.modelLoader.modelLoaded.connect(self.onModelsLoaded)
        self.modelLoader.start()

    def onModelsLoaded(self, sp, facerec):
        self.sp = sp
        self.facerec = facerec

    def show_error_message(self, message):
        QtWidgets.QMessageBox.critical(self, "Error", message)

    def trigger_warning_sound(self, check, num):
        if num == 1:
            self.current_page.warning_sound.play()
        elif num == 2:
            self.current_page.warning_sound_2.play()




    def layoutIn1(self):
        self.buttonGroupBox.setVisible(False)
        self.buttonGroupBox_2.setVisible(True)

    def layoutIn2(self):
        self.buttonGroupBox.setVisible(True)
        self.buttonGroupBox_2.setVisible(False)


    def resizeEvent(self, event):
        self.bottom_right_resizer.setGeometry(self.width() - 16, self.height() - 16, 16, 16)

    def mousePressEvent(self, event):
        if event.button() == QtCore.Qt.LeftButton:
            self.__isDragging = True
            self.__startPos = event.pos()
            if self.bottom_right_resizer.rect().contains(event.pos()):
                self.start_resizing(event)

    def mouseReleaseEvent(self, event):
        if event.button() == QtCore.Qt.LeftButton:
            self.__isDragging = False
            if hasattr(self, 'resize_origin'):
                del self.resize_origin

    def mouseMoveEvent(self, event):
        if self.__isDragging:
            delta = event.pos() - self.__startPos
            self.move(self.pos() + delta)
        if hasattr(self, 'resize_origin'):
            delta = event.pos() - self.resize_origin
            new_size = self.resize_origin_size + delta
            self.resize(new_size.width(), new_size.height())

    def start_resizing(self, event):
        self.resize_origin = event.pos()
        self.resize_origin_size = QtCore.QSize(self.width(), self.height())

    def mouseDoubleClickEvent(self, event):
        if event.button() == Qt.LeftButton:
            if self.isMaximized():
                self.showNormal()
            else:
                self.showMaximized()

    def enlarge(self):
        if self.isMaximized():
            self.showNormal()
        else:
            self.showMaximized()

    def backToLoginForm(self):
        self.close()
        if hasattr(self.current_page, "timer"):
            self.current_page.timer.stop()
            self.current_page.cam.release()
        widget.show()

    def on_webcam_initializedCollect(self, success):
        if success:
            self.current_page.cap = self.current_page.thread.cap
        else:
            print("Webcam initialization failed")
        self.current_page.timer = QtCore.QTimer(self)
        self.current_page.stop_button.setEnabled(True)
        self.current_page.timer.timeout.connect(self.update_webcam_collect)
        self.current_page.minW = 0
        self.current_page.minH = 0
        self.current_page.timer.start(30)

    def on_webcam_initializedRecog(self, success):
        if success:
            self.current_page.cam = self.current_page.webcam_thread_recog.cam
        else:
            print("Webcam initialization failed")
        self.current_page.timer = QtCore.QTimer(self)
        self.current_page.stopRecordBtn.setEnabled(True)
        self.current_page.cam.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
        self.current_page.cam.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)
        self.current_page.minW = int(0.1 * self.current_page.cam.get(3))
        self.current_page.minH = int(0.1 * self.current_page.cam.get(4))
        self.current_page.timer.timeout.connect(self.update_frame_recog)
        self.current_page.timer.start(30)

    def show_CollectData(self):
        if self.openRecognitionCamBtn.setDisabled(False):
            self.gotoFirstMenu_recog()
        self.openRecognitionCamBtn.setDisabled(False)
        self.loadExcelButton.setDisabled(False)
        self.openRecognitionCamBtn_3.setDisabled(False)
        self.loadExcelButton_3.setDisabled(False)
        self.collectDataBtn.setDisabled(True)
        self.collectDataBtn_3.setDisabled(True)
        self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
        self.setMinimumSize(0, 0)
        self.resize(1050, 800)
        if self.current_page:
            self.current_page.hide()
        self.current_page = loadUi('UI/Collect_Data.ui')
        self.content_layout.addWidget(self.current_page)

        with open("Styles/recorderStyle.qss", "r") as f:
            self.stylesheet = f.read()
        self.current_page.file = QFile("Styles/recorderStyle.qss")
        self.current_page.file.open(QFile.ReadOnly | QFile.Text)
        self.current_page.stream = QTextStream(self.current_page.file)
        self.current_page.stylesheet = self.current_page.stream.readAll()
        self.current_page.setStyleSheet(self.current_page.stylesheet)

        # Initialize MTCNN detector
        self.current_page.detector = MTCNN()
        self.current_page.counter = 0
        self.current_page.all_ids = set()

        # Retrieving all the prefixes of the images in the current directory
        for file in os.listdir(os.path.join(os.getcwd(), "dataset")):
            if file.endswith(".jpg"):
                self.current_page.all_ids.add(file.split("_")[0])

        # Connect signals
        self.current_page.start_button.clicked.connect(self.start_webcam_collect)
        self.current_page.stop_button.clicked.connect(self.stop_webcam_collect)
        self.current_page.backBtn.clicked.connect(self.gotoFirstMenu_collect)
        self.current_page.trainBtn.clicked.connect(self.start_Train)
        self.current_page.start_button.setIconSize(QSize(int(self.current_page.start_button.width() / 100 * 50),
                                                      int(self.current_page.start_button.height() / 100 * 50)))
        self.current_page.stop_button.setIconSize(QSize(int(self.current_page.stop_button.width() / 100 * 50),
                                                          int(self.current_page.stop_button.height() / 100 * 50)))
        self.current_page.trainBtn.setIconSize(QSize(int(self.current_page.trainBtn.width() / 100 * 50),
                                                     int(self.current_page.trainBtn.height() / 100 * 50)))
        self.current_page.backBtn.setIconSize(QSize(int(self.current_page.backBtn.width() / 100 * 50),
                                                    int(self.current_page.backBtn.height() / 100 * 50)))
        self.current_page.start_button.setCursor(Qt.PointingHandCursor)
        self.current_page.stop_button.setCursor(Qt.PointingHandCursor)
        self.current_page.backBtn.setCursor(Qt.PointingHandCursor)
        self.current_page.trainBtn.setCursor(Qt.PointingHandCursor)
        self.current_page.start_button.setToolTip("Start Recording")
        self.current_page.stop_button.setToolTip("Stop Recording")
        self.current_page.trainBtn.setToolTip("Train Model")
        self.current_page.backBtn.setVisible(False)
        self.current_page.stop_button.setEnabled(False)
        self.current_page.trainBtn.setEnabled(True)

    def start_Train(self):
        self.setDisabled(True)
        trainer = Train_MainWindow(self)
        trainer.exec()

    def stop_webcam_collect(self):
        # Release the VideoCapture object
        self.current_page.cap.release()

        # Stop the QTimer
        self.current_page.timer.stop()

        # Enable buttons and inputs
        self.openRecognitionCamBtn.setEnabled(True)
        self.loadExcelButton.setEnabled(True)
        self.current_page.id_edit.setEnabled(True)
        self.current_page.start_button.setEnabled(True)
        self.current_page.stop_button.setEnabled(False)
        self.current_page.trainBtn.setEnabled(True)
        self.current_page.backBtn.setEnabled(True)

        # Reset the minimum size of the window
        self.setMinimumSize(1050, 800)

        # Reset the size policy of the window
        self.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.MinimumExpanding)

        # Resize the window
        self.resize(1175, 830)

    def gotoFirstMenu_collect(self):
        self.setMinimumSize(1050, 800)
        self.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.MinimumExpanding)
        self.resize(1050, 800)
        self.collectDataBtn.setDisabled(False)
        if self.current_page:
            self.current_page.hide()
        for i in reversed(range(self.content_layout.count())):
            self.content_layout.itemAt(i).widget().setParent(None)
        self.current_page = loadUi('UI/empty_menu.ui')
        self.content_layout.addWidget(self.current_page)

    def start_webcam_collect(self):
        check = self.test_cam(1)
        if check == 1:
            QMessageBox.warning(None, "Warning", "Webcam not found.")
            return
        id = self.current_page.id_edit.text()
        # validating the input only contains numbers
        if id.isnumeric():
            if id in self.current_page.all_ids:
                app = QApplication.instance()
                if app is None:
                    app = QApplication(sys.argv)
                # Load stylesheet
                with open('Styles/messageBox.qss') as f:
                    style = f.read()
                    QApplication.setStyle('Fusion')
                    app.setStyleSheet(style)
                msgBox = QMessageBox()
                msgBox.setIcon(QMessageBox.Warning)
                msgBox.setText("This ID already exists, please try another one")
                msgBox.setWindowTitle("Warning")
                msgBox.setStandardButtons(QMessageBox.Ok)
                msgBox.exec_()
                print("This ID already exists, please try another one")
                return
            else:
                self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
                self.setMinimumSize(0, 0)
                self.current_page.id = id
                self.current_page.id_edit.setEnabled(False)
                self.current_page.start_button.setEnabled(False)
                self.current_page.backBtn.setEnabled(False)
                self.openRecognitionCamBtn.setEnabled(False)
                self.current_page.trainBtn.setEnabled(False)
                self.loadExcelButton.setEnabled(False)
                # def initialize_webcam():
                #     self.current_page.cap = cv2.VideoCapture(0)
                #     self.current_page.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
                #     self.current_page.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)
                #
                # Thread(target=initialize_webcam).start()
                self.current_page.thread = WebcamThreadCollect(self)
                self.current_page.thread.initialized.connect(lambda success: self.on_webcam_initializedCollect(success))
                self.current_page.thread.start()
        elif id == "":
            msgBox = QMessageBox()
            msgBox.setIcon(QMessageBox.Warning)
            msgBox.setText("Please enter an id")
            msgBox.setWindowTitle("Warning")
            msgBox.setStandardButtons(QMessageBox.Ok)
            msgBox.exec_()
        else:
            app = QApplication.instance()
            if app is None:
                app = QApplication(sys.argv)
            with open('Styles/messageBox.qss') as f:
                style = f.read()
                QApplication.setStyle('Fusion')
                app.setStyleSheet(style)
            msgBox = QMessageBox()
            msgBox.setIcon(QMessageBox.Warning)
            msgBox.setText("Invalid input, Please enter only numbers")
            msgBox.setWindowTitle("Warning")
            msgBox.setStandardButtons(QMessageBox.Ok)
            msgBox.exec_()

    def update_webcam_collect(self):
        if self.current_page.cap is None:
            return
        ret, frame = self.current_page.cap.read()
        if frame is None:
            return
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        if ret:
            # Detect faces in the frame
            faces = self.current_page.detector.detect_faces(frame)

            # Draw bounding boxes around the faces
            for face in faces:
                x, y, width, height = face['box']
                cv2.rectangle(frame, (x, y), (x + width, y + height), (0, 0, 255), 2)

                # Crop the face from the frame
                face_img = frame[y:y + height, x:x + width]

                # Convert the face to grayscale
                gray_face = cv2.cvtColor(face_img, cv2.COLOR_BGR2RGB)

                # Save the aligned gray face to a file using the id as prefix
                dataset_path = os.path.join(os.getcwd(), "dataset")
                if not os.path.exists(dataset_path):
                    os.mkdir(dataset_path)
                cv2.imwrite(os.path.join(dataset_path, self.current_page.id + "_face" + str(self.current_page.counter) + ".jpg"), gray_face)
                self.current_page.all_ids.add(self.current_page.id)
                self.current_page.counter += 1
                if self.current_page.counter >= 100:
                    self.stop_webcam_collect()
                    self.current_page.counter = 0
                    self.current_page.id = None
                    reply = QMessageBox.question(self, 'Message',
                                                 "Do you want to start the model training process?", QMessageBox.Yes |
                                                 QMessageBox.No, QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        self.setDisabled(True)
                        trainer = Train_MainWindow(self)
                        trainer.exec()
                    else:
                        pass
                    break

                # Display the frame with bounding boxes
            # self.current_page.webcam_label.setPixmap(QtGui.QPixmap.fromImage(
            #     QtGui.QImage(frame, frame.shape[1], frame.shape[0], QtGui.QImage.Format_RGB888)))

            h, w, ch = frame.shape
            bytes_per_line = ch * w
            # convert_to_Qt_format = QtGui.QImage(self.current_page.img.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
            # p = convert_to_Qt_format.scaled(900, 700)
            self.current_page.webcam_label.setPixmap(QtGui.QPixmap.fromImage(
                QtGui.QImage(frame.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)))

    def show_recognizer(self):
        self.loadExcelButton.setDisabled(False)
        self.collectDataBtn.setDisabled(False)
        self.loadExcelButton_3.setDisabled(False)
        self.collectDataBtn_3.setDisabled(False)
        self.openRecognitionCamBtn.setDisabled(True)
        self.openRecognitionCamBtn_3.setDisabled(True)
        self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
        self.setMinimumSize(0, 0)
        self.resize(1050, 800)
        if self.current_page:
            self.current_page.hide()
        self.current_page = loadUi('UI/Recognizer.ui')
        self.content_layout.addWidget(self.current_page)

        with open("Styles/recognizerStyle.qss", "r") as f:
            self.stylesheet = f.read()
        self.current_page.file = QFile("Styles/recognizerStyle.qss")
        self.current_page.file.open(QFile.ReadOnly | QFile.Text)
        self.current_page.stream = QTextStream(self.current_page.file)
        self.current_page.stylesheet = self.current_page.stream.readAll()
        self.current_page.setStyleSheet(self.current_page.stylesheet)
        self.current_page.webcam_thread_recog = None
        self.current_page.scrollArea.setVisible(False)
        self.current_page.unknown_faces_label.setVisible(False)
        self.current_page.unknown_faces_frame.setVisible(False)
        self.current_page.work_dir = os.getcwd()
        self.current_page.recognizer = cv2.face.LBPHFaceRecognizer_create()
        self.current_page.recognizer.read(self.current_page.work_dir + '/trainer/trainer.yml')
        self.current_page.work_dir = os.getcwd()
        self.current_page.warning_sound = QSound(self.current_page.work_dir + '/Sound/warning.wav')
        self.current_page.warning_sound_2 = QSound(self.current_page.work_dir + '/Sound/warning-sound-2.wav')

        self.current_page.faceCascade = cv2.CascadeClassifier(
            os.path.join(self.current_page.work_dir, "haarcascade_frontalface_default.xml"))
        self.current_page.font = cv2.FONT_HERSHEY_SIMPLEX

        # Define min window size to be recognized as a face
        self.current_page.backBtn.clicked.connect(self.gotoFirstMenu_recog)
        self.current_page.recordBtn.clicked.connect(self.start_webcam_recog)
        self.current_page.stopRecordBtn.clicked.connect(self.stop_webcam_recog)
        self.current_page.clearImage.clicked.connect(self.clear_images)
        self.current_page.backBtn.setIconSize(QSize(int(self.current_page.backBtn.width() / 100 * 50),
                                                         int(self.current_page.backBtn.height() / 100 * 50)))
        self.current_page.recordBtn.setIconSize(QSize(int(self.current_page.recordBtn.width() / 100 * 50),
                                                        int(self.current_page.recordBtn.height() / 100 * 50)))
        self.current_page.stopRecordBtn.setIconSize(QSize(int(self.current_page.stopRecordBtn.width() / 100 * 50),
                                                     int(self.current_page.stopRecordBtn.height() / 100 * 50)))
        self.current_page.backBtn.setVisible(False)
        self.current_page.clearImage.setVisible(False)
        self.current_page.recordBtn.setCursor(Qt.PointingHandCursor)
        self.current_page.stopRecordBtn.setCursor(Qt.PointingHandCursor)
        self.current_page.backBtn.setCursor(Qt.PointingHandCursor)
        self.current_page.recordBtn.setToolTip("Start Recording")
        self.current_page.stopRecordBtn.setToolTip("Stop Recording")
        self.current_page.stopRecordBtn.setEnabled(False)
        self.current_page.last_unknown_face = None
        self.current_page.last_unknown_face_2 = None
        self.current_page.last_unknown_face_time = datetime.now()
        self.current_page.last_unknown_face_time_2 = datetime.now()


    def add_unknown_face_to_layout(self, img_label):
        unknown_label = QtWidgets.QLabel()
        unknown_label.setPixmap(QtGui.QPixmap.fromImage(img_label))
        self.current_page.unknown_faces_layout.addWidget(unknown_label)

    def clear_images(self):
        for i in reversed(range(self.current_page.unknown_faces_layout.count())):
            widget = self.current_page.unknown_faces_layout.itemAt(i).widget()
            if widget is not None:
                widget.setParent(None)

    def gotoFirstMenu_recog(self):
        self.current_page.webcam_thread_recog.quit()
        self.current_page.webcam_thread_recog.wait()
        if self.openRecognitionCamBtn.setDisabled(True):
            if self.current_page.timer and self.current_page.cam:
                self.stop_webcam_recog()
        self.openRecognitionCamBtn.setDisabled(False)
        self.loadExcelButton.setEnabled(True)
        self.collectDataBtn.setEnabled(True)
        self.setMinimumSize(1050, 800)
        self.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.MinimumExpanding)
        self.resize(1050, 800)
        if hasattr(self.current_page, "webcam_thread_recog"):
            if self.current_page.webcam_thread_recog.isRunning():
                self.current_page.webcam_thread_recog.stop_signal.emit()
        if self.current_page:
            self.current_page.hide()
        for i in reversed(range(self.content_layout.count())):
            self.content_layout.itemAt(i).widget().setParent(None)
        self.current_page = loadUi('UI/empty_menu.ui')
        self.content_layout.addWidget(self.current_page)

    def test_cam(self, source):
        self.cap = cv2.VideoCapture(source)
        if not self.cap.isOpened():
            self.cap.release()
            return 1
        return 0

    def start_webcam_recog(self):
        check = self.test_cam(1)
        if check == 1:
            QMessageBox.warning(None, "Warning", "Webcam not found.")
            return
        else:
            self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
            self.setMinimumSize(0, 0)
            self.current_page.recordBtn.setEnabled(False)
            self.current_page.imageLabel.setVisible(True)
            if self.current_page.unknown_faces_layout.count() > 0:
                self.current_page.scrollArea.setVisible(True)
                self.current_page.unknown_faces_label.setVisible(True)
                self.current_page.clearImage.setVisible(True)
                self.current_page.unknown_faces_frame.setVisible(True)
            self.loadExcelButton.setEnabled(False)
            self.collectDataBtn.setEnabled(False)
            self.current_page.webcam_thread_recog = WebcamThreadRecog(self)
            self.current_page.webcam_thread_recog.initialized.connect(
                lambda success: self.on_webcam_initializedRecog(success))
            self.current_page.webcam_thread_recog.start()

    def stop_webcam_recog(self):
        self.resize(1200, 850)
        self.setMinimumSize(1200, 830)
        self.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.MinimumExpanding)
        self.loadExcelButton.setEnabled(True)
        self.collectDataBtn.setEnabled(True)
        self.current_page.recordBtn.setEnabled(True)
        self.current_page.stopRecordBtn.setEnabled(False)
        self.current_page.scrollArea.setVisible(False)
        self.current_page.unknown_faces_label.setVisible(False)
        self.current_page.clearImage.setVisible(False)
        self.current_page.unknown_faces_frame.setVisible(False)
        self.current_page.timer.stop()
        self.current_page.cam.release()
        self.current_page.imageLabel.setPixmap(QtGui.QPixmap())





    def handle_recognition_result(self, result):
        if 'face' not in result:
            img = result['img']
            h, w, ch = img.shape
            bytes_per_line = ch * w
            self.current_page.imageLabel.setPixmap(QtGui.QPixmap.fromImage(
                QtGui.QImage(img.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)))

            self.current_page.unknown_faces_layout.update()
            self.update_frame_recog()  # update the image in the GUI
            return

        img = result['img']
        h, w, ch = img.shape
        bytes_per_line = ch * w
        self.current_page.imageLabel.setPixmap(QtGui.QPixmap.fromImage(
            QtGui.QImage(img.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)))

        self.current_page.unknown_faces_layout.update()
        self.update_frame_recog()  # update the image in the GUI

    def set_ScrollArea_Visible(self):
        self.current_page.scrollArea.setVisible(True)
        self.current_page.unknown_faces_label.setVisible(True)
        self.current_page.clearImage.setVisible(True)
        self.current_page.unknown_faces_frame.setVisible(True)

    def clean_CamRecog(self):
        self.current_page.imageLabel.setPixmap(QtGui.QPixmap())

    def update_frame_recog(self):
        if hasattr(self.current_page, 'recognition_thread'):
            if self.current_page.recognition_thread is not None and self.current_page.recognition_thread.isRunning():
                return
        ret, img = self.current_page.cam.read()
        if not hasattr(self.current_page, 'cam') or img is None:
            return
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        self.current_page.color_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        self.current_page.threshold = 65

        # Load the saved face encodings
        with open(self.current_page.work_dir + '/encodings.pkl', 'rb') as f:
            saved_encodings = pickle.load(f)
        embeddings = saved_encodings["encodings"]
        self.current_page.ids = saved_encodings["ids"]
        self.current_page.recognition_thread = FaceRecognitionThread(embeddings, self.current_page.ids, img,
                                                                     self.current_page.threshold,
                                                                     self.current_page.work_dir, self.sp, self.facerec)
        self.current_page.recognition_thread.recognition_complete.connect(self.handle_recognition_result)
        self.current_page.recognition_thread.add_image.connect(self.add_unknown_face_to_layout)
        self.current_page.recognition_thread.display_scrollArea.connect(self.set_ScrollArea_Visible)
        self.current_page.recognition_thread.trigger_sound.connect(self.trigger_warning_sound)
        self.current_page.recognition_thread.start()



    def show_Load_Excel(self):
        if os.path.isfile("face_log.xlsx"):
            self.openRecognitionCamBtn.setDisabled(False)
            self.collectDataBtn.setDisabled(False)
            self.loadExcelButton.setDisabled(True)
            self.openRecognitionCamBtn_3.setDisabled(False)
            self.collectDataBtn_3.setDisabled(False)
            self.loadExcelButton_3.setDisabled(True)
            self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
            self.setMinimumSize(1050, 800)
            if self.current_page:
                self.current_page.hide()
            self.current_page = loadUi('UI/table.ui')
            self.content_layout.addWidget(self.current_page)
            self.load_excel()
        else:
            QMessageBox.warning(None, "Warning", "Log does not exist yet")
            return

    def load_excel(self):
        # Show a file dialog to choose an Excel file
        #filename, _ = QFileDialog.getOpenFileName(self, "Open Excel File", os.getcwd()+"\Log", "Excel Files (*.xlsx *.xls)")
        filename = os.getcwd() + "/face_log.xlsx"
        if filename:
            # Load the data from the Excel file into a pandas DataFrame
            df = pd.read_excel(filename)

            # Create a new QStandardItemModel and set it to the QTableView
            model = QStandardItemModel(df.shape[0], df.shape[1], self.current_page)
            self.current_page.table.setModel(model)

            # Add the column titles to the model
            for column in range(df.shape[1]):
                title = str(df.columns[column])
                item = QStandardItem(title)
                model.setHorizontalHeaderItem(column, item)

            # Add the data to the model
            for row in range(df.shape[0]):
                for column in range(df.shape[1]):
                    item = QStandardItem(str(df.iloc[row, column]))
                    model.setItem(row, column, item)

class Train_MainWindow(QtWidgets.QDialog):
    message_signal = QtCore.pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)

        uic.loadUi("UI/Train_model.ui", self)

        with open("Styles/style.qss", "r") as f:
            self.stylesheet = f.read()
        self.file = QFile("Styles/trainerStyle.qss")
        self.file.open(QFile.ReadOnly | QFile.Text)
        self.stream = QTextStream(self.file)
        self.stylesheet = self.stream.readAll()
        self.setStyleSheet(self.stylesheet)

        self.backBtn.clicked.connect(self.gotoFirstMenu)
        self.stop_training = False
        self.work_dir = os.getcwd()
        self.path = self.work_dir + '/dataset'
        self.progress_bar.setVisible(False)
        self.message_signal.connect(self.final_message_signal)

        self.show()

        # Wait for 3 seconds before starting training
        self.start_training_timer = QtCore.QTimer()
        self.start_training_timer.timeout.connect(self.start_training)
        self.start_training_timer.start(3000)
        self.label.setText("Waiting for the Training Process to Start")

    def start_training(self):
        self.start_training_timer.stop()
        self.label.setText("Training is in progress")
        self.train()

    def gotoFirstMenu(self):
        self.stop_training = True
        if self.start_training_timer.remainingTime() == 0:
            self.thread.join()
        parentWidget = self.parent()
        parentWidget.setDisabled(False)
        self.accept()

    def train(self):
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.timer = QTimer(self, timeout=self.onTimeout)
        self.timer.start(randint(1, 3) * 1000)
        self.stop_training = False  # reset the stop_training flag
        self.thread = Thread(target=self.train_thread)
        self.thread.start()

    def train_thread(self):
        detector = dlib.get_frontal_face_detector()
        sp = dlib.shape_predictor(self.work_dir+"/trainer/shape_predictor_68_face_landmarks.dat")
        facerec = dlib.face_recognition_model_v1(self.work_dir+"/trainer/dlib_face_recognition_resnet_model_v1.dat")

        def get_images_and_labels(path):
            image_paths = [os.path.join(path, f) for f in os.listdir(path)]
            face_encodings = []
            ids = []
            for image_path in image_paths:
                id = int(os.path.split(image_path)[-1].split("_")[0])
                img = dlib.load_rgb_image(image_path)
                dets = detector(img, 1)
                for det in dets:
                    shape = sp(img, det)
                    face_encoding = facerec.compute_face_descriptor(img, shape)
                    face_encodings.append(face_encoding)
                    ids.append(id)
            return face_encodings, ids

        face_encodings, ids = get_images_and_labels(self.path)
        if self.stop_training:
            parentWidget = self.parent()
            parentWidget.setDisabled(False)
            return
        with open(self.work_dir + '/encodings.pkl', 'wb') as f:
            pickle.dump({"encodings": face_encodings, "ids": ids}, f)
        self.progress_bar.setVisible(False)
        self.message_signal.emit()


    def final_message_signal(self):
        reply = QMessageBox.information(self, "Training results", "Training completed")
        if reply == QMessageBox.Ok:
            parentWidget = self.parent()
            parentWidget.setDisabled(False)
            self.accept()

    def onTimeout(self):
        if self.progress_bar.value() >= 100:
            self.timer.stop()
            self.timer.deleteLater()
            del self.timer
            return
        self.progress_bar.setValue(self.progress_bar.value() + 1)

if __name__ == '__main__':
    # Create application
    app = QApplication(sys.argv)

    # Create object
    welcome = LoginForm()
    # Working with multiple forms/screens (swich screen).
    widget = QtWidgets.QStackedWidget()
    # Add widget
    widget.addWidget(welcome)
    widget.setFixedHeight(345)
    widget.setFixedWidth(415)
    # widget.setWindowOpacity(0.5)
    widget.setWindowTitle('Security System using Face Recognition')  # set window title in Python PyQt5
    widget.setWindowIcon(QtGui.QIcon('./icons/logo.png'))
    widget.show()

    try:
        sys.exit(app.exec_())
    except:
        print("Exiting")